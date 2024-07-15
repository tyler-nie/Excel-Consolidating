from openpyxl import load_workbook
from config import source_file, master_list

source_workbook = load_workbook(source_file)
source_sheet = source_workbook.active

destination_workbook = load_workbook(master_list)
destination_sheet = destination_workbook.active

seen_sources = []


def main():
    get_data()


def get_data():
    """
    Gets data and pushes through to append to master list
    """
    column_index = 1
    count = 0
    results = []
    folder_num = ""

    # Iterate through the cells in the column A
    for row in source_sheet.iter_rows(min_col=column_index):
        # Check if the cell is empty or contains headers
        if (row[0].value is None) or (row[0].font and row[0].font.bold and count >= 1):
            continue  # Skip empty cells and headers

        # Check if font is bold
        if row[0].font and row[0].font.bold and count == 0:
            count = 1
            folder_num = get_folder_num(row[0].value)

        else:
            results.append(get_results(row))
    push_data(folder_num, results)
    results = []  # Clear results after pushing


def get_folder_num(row):
    """
    Gets the folder number
    """
    folder_index = row.find("FOLDER NUM.:")

    if folder_index != -1:
        # Extract Folder Num from String
        folder_num_str = row[folder_index + len("FOLDER NUM.:"):]
        folder_num = folder_num_str.split()[0]

        print(f"Folder Number: {folder_num}")
        return folder_num
    else:
        print("Folder Number not found in input string.")


def get_results(row):
    """
    Gets the data required from the test result
    """
    source = row[0].value
    date = (str(row[3].value)).split()[0]
    method = row[8].value
    analyte = row[9].value
    result = row[10].value
    return {
        "source": source,
        "date": date,
        "analyte": analyte,
        "method": method,
        "result": result,
    }


def find_column_index(result):
    """
    Gets the index of the specified results analyte
    """
    for row in destination_sheet.iter_rows(min_row=1, max_row=1, values_only=True):
        for index, cell_value in enumerate(row, start=1):
            if cell_value == result["analyte"]:
                return index
    return None


def create_new_column(analyte, method):
    """
    Creates new column with the new analyte
    """
    index = destination_sheet.max_column + 1
    destination_sheet.cell(row=1, column=index, value=analyte)
    destination_sheet.cell(row=2, column=index, value=method)
    print(f"Created new column at index {index} for analyte {analyte} with method {method}")
    return index


def get_row_number(source):
    """
    Gets the row number of a seen source
    """
    for seen_source, row_number in seen_sources:
        if seen_source == source:
            return row_number
    return None


def push_data(folder_num, data):
    """
    Pushes data to the masterlist
    """
    for result in data:
        source = result["source"]
        index = find_column_index(result)
        row_num = get_row_number(source)

        if row_num is None:
            row_num = destination_sheet.max_row + 1
            seen_sources.append((source, row_num))

        if index is None:  # If the index is none, analyte does not exists in masterlist
            index = create_new_column(result["analyte"], result["method"])

        # print(f"Writing data at row {row_num}, column {index}")  # Troubleshooting
        
        destination_sheet.cell(row=row_num, column=1, value=result["source"])
        destination_sheet.cell(row=row_num, column=2, value=result["date"])
        destination_sheet.cell(row=row_num, column=3, value=folder_num)
        destination_sheet.cell(row=row_num, column=index, value=result["result"])

    destination_workbook.save(master_list)


if __name__ == "__main__":
    main()

source_workbook.close()
destination_workbook.close()
